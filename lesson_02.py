"""
============================
Project: Python-class
Author:柠檬班-dd
Time:2022/9/9 17:45
E-mail:1161340814@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""

import requests
from openpyxl import load_workbook
# #读取数据函数
def duqushuju(biao,biaodan):
    wb=load_workbook(biao)
    sheet=wb[biaodan]
    max_row=sheet.max_row
    case_list=[]
    for i in range(2,max_row+1):
        case=dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,
        data=sheet.cell(row=i,column=6).value,
        expected=sheet.cell(row=i,column=7).value)
        case_list.append(case)
    return case_list

# #发送请求函数
def fasongqingqiu(url,data):
    header={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    respoonse=requests.post(url=url,json=data,headers=header)
    return (respoonse.json())

#回写数据函数
def huixieshuju(biao,biaodan,row,column,shijijieguo):
    wb=load_workbook(biao)
    sheet=wb[biaodan]
    sheet.cell(row=row,column=column).value=shijijieguo
    wb.save(biao)

# #执行用例函数
def zhixingyongli(biao,biaodan):
    cases=duqushuju(biao,biaodan)
    for case in cases:
        case_id=case.get("case_id")
        url=case["url"]
        data=case["data"]
        data=eval(data)
        qiwangjieguo=case["expected"]
        qiwangjieguo=eval(qiwangjieguo)
        shijijieguo=fasongqingqiu(url=url,data=data)
        qiwangjieguo_msg=qiwangjieguo["msg"]
        shijijieguo_msg=shijijieguo["msg"]
        print("期望结果是:{}".format(qiwangjieguo_msg))
        print("实际结果是:{}".format(shijijieguo_msg))
        if shijijieguo_msg==qiwangjieguo_msg:
            print("第{}条用例执行通过".format(case_id))
            zuizhongjieguo="通过"
        else:
            print("第{}条用例执行不通过".format(case_id))
            zuizhongjieguo="不通过"
        huixieshuju(biao,biaodan,case_id+1,8,zuizhongjieguo)

zhixingyongli("testcase_api_wuye.xlsx","register")


